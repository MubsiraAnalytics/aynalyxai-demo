# AynalyxAI Demo - Streamlit Cloud Version
# Sample data only - no file upload

import streamlit as st
import pandas as pd
import numpy as np
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import StandardScaler
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# Page config - CENTERED layout for mobile-friendly single page
st.set_page_config(
    page_title="AynalyxAI Demo",
    page_icon="🔬",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# Custom CSS for single-page mobile-friendly design (NO SIDEBAR)
st.markdown("""
<style>
    /* HIDE SIDEBAR COMPLETELY */
    [data-testid="stSidebar"] {
        display: none !important;
    }
    [data-testid="stSidebarCollapsedControl"] {
        display: none !important;
    }
    
    /* Mobile-first responsive container */
    .block-container {
        padding: 1rem 1rem 3rem 1rem !important;
        max-width: 100% !important;
    }
    
    @media (min-width: 768px) {
        .block-container {
            padding: 2rem 2rem 4rem 2rem !important;
            max-width: 900px !important;
            margin: 0 auto;
        }
    }
    
    /* Main header styling - MOBILE RESPONSIVE */
    .main-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.5rem 1rem;
        border-radius: 16px;
        color: white;
        text-align: center;
        margin-bottom: 1rem;
        box-shadow: 0 8px 30px rgba(102, 126, 234, 0.3);
    }
    
    @media (min-width: 768px) {
        .main-header {
            padding: 2.5rem 2rem;
            border-radius: 20px;
            margin-bottom: 1.5rem;
        }
    }
    
    .main-header h1 {
        margin: 0;
        font-size: 1.8rem;
        font-weight: 800;
        text-shadow: 0 2px 10px rgba(0,0,0,0.2);
    }
    
    @media (min-width: 768px) {
        .main-header h1 {
            font-size: 2.5rem;
        }
    }
    
    .main-header p {
        margin: 0.5rem 0 0 0;
        opacity: 0.95;
        font-size: 0.95rem;
    }
    
    @media (min-width: 768px) {
        .main-header p {
            font-size: 1.15rem;
            margin: 0.75rem 0 0 0;
        }
    }
    
    /* Demo notice - MOBILE RESPONSIVE */
    .demo-badge {
        background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%);
        color: #92400e;
        padding: 0.75rem 1rem;
        border-radius: 10px;
        font-weight: 700;
        display: block;
        text-align: center;
        margin-bottom: 1rem;
        border: 2px solid #f59e0b;
        font-size: 0.85rem;
        box-shadow: 0 4px 15px rgba(245, 158, 11, 0.2);
    }
    
    @media (min-width: 768px) {
        .demo-badge {
            padding: 1rem 1.5rem;
            font-size: 1rem;
            margin-bottom: 1.5rem;
        }
    }
    
    /* Sample buttons grid - MOBILE 2x2 */
    .sample-btn-grid {
        display: grid;
        grid-template-columns: repeat(2, 1fr);
        gap: 0.5rem;
        margin: 1rem 0;
    }
    
    @media (min-width: 768px) {
        .sample-btn-grid {
            grid-template-columns: repeat(4, 1fr);
            gap: 0.75rem;
        }
    }
    
    /* Anomaly level colors - matching real app */
    .anomaly-critical {
        background-color: #dc2626 !important;
        color: white !important;
        padding: 2px 8px;
        border-radius: 4px;
        font-weight: bold;
    }
    
    .anomaly-high {
        background-color: #ef4444 !important;
        color: white !important;
        padding: 2px 8px;
        border-radius: 4px;
        font-weight: bold;
    }
    
    .anomaly-medium {
        background-color: #f59e0b !important;
        color: white !important;
        padding: 2px 8px;
        border-radius: 4px;
        font-weight: bold;
    }
    
    .anomaly-low {
        background-color: #22c55e !important;
        color: white !important;
        padding: 2px 8px;
        border-radius: 4px;
        font-weight: bold;
    }
    
    /* Results table styling */
    .results-container {
        background: #f8fafc;
        border-radius: 12px;
        padding: 1.5rem;
        border: 1px solid #e2e8f0;
    }
    
    /* Stats cards - MOBILE RESPONSIVE */
    .stat-card {
        background: white;
        padding: 0.75rem 0.5rem;
        border-radius: 12px;
        text-align: center;
        box-shadow: 0 2px 10px rgba(0,0,0,0.06);
        border-left: 4px solid;
    }
    
    @media (min-width: 768px) {
        .stat-card {
            padding: 1.25rem 1rem;
            border-radius: 16px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.08);
        }
    }
    
    .stat-card .stat-number {
        font-size: 1.5rem;
        font-weight: 800;
        line-height: 1;
    }
    
    @media (min-width: 768px) {
        .stat-card .stat-number {
            font-size: 2.2rem;
        }
    }
    
    .stat-card .stat-label {
        font-size: 0.65rem;
        color: #666;
        margin-top: 0.25rem;
        font-weight: 600;
    }
    
    @media (min-width: 768px) {
        .stat-card .stat-label {
            font-size: 0.8rem;
            margin-top: 0.5rem;
        }
    }
    
    .stat-critical { border-color: #dc2626; background: linear-gradient(135deg, #fff 0%, #fef2f2 100%); }
    .stat-high { border-color: #ef4444; background: linear-gradient(135deg, #fff 0%, #fef2f2 100%); }
    .stat-medium { border-color: #f59e0b; background: linear-gradient(135deg, #fff 0%, #fffbeb 100%); }
    .stat-low { border-color: #22c55e; background: linear-gradient(135deg, #fff 0%, #f0fdf4 100%); }
    
    /* How it works section - MOBILE RESPONSIVE */
    .how-it-works {
        background: linear-gradient(135deg, #f8fafc 0%, #e0f2fe 100%);
        border-radius: 12px;
        padding: 1rem;
        margin-bottom: 1rem;
        border: 1px solid #bae6fd;
    }
    
    @media (min-width: 768px) {
        .how-it-works {
            padding: 1.5rem 2rem;
            border-radius: 16px;
            margin-bottom: 1.5rem;
        }
    }
    
    .how-it-works h3 {
        color: #0c4a6e;
        margin: 0 0 0.75rem 0;
        font-size: 0.95rem;
    }
    
    @media (min-width: 768px) {
        .how-it-works h3 {
            font-size: 1.1rem;
            margin: 0 0 1rem 0;
        }
    }
    
    .steps-container {
        display: grid;
        grid-template-columns: repeat(3, 1fr);
        gap: 0.5rem;
    }
    
    @media (min-width: 768px) {
        .steps-container {
            grid-template-columns: repeat(5, 1fr);
            gap: 1rem;
        }
    }
    
    .step-box {
        text-align: center;
        padding: 0.5rem;
        background: white;
        border-radius: 8px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.05);
    }
    
    .step-box .step-num {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        width: 24px;
        height: 24px;
        border-radius: 50%;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        font-weight: bold;
        font-size: 0.75rem;
        margin-bottom: 0.25rem;
    }
    
    @media (min-width: 768px) {
        .step-box .step-num {
            width: 32px;
            height: 32px;
            font-size: 0.9rem;
            margin-bottom: 0.5rem;
        }
    }
    
    .step-box .step-text {
        color: #475569;
        font-size: 0.65rem;
        line-height: 1.2;
    }
    
    @media (min-width: 768px) {
        .step-box .step-text {
            font-size: 0.85rem;
            line-height: 1.3;
        }
    }
    
    /* Info card - MOBILE RESPONSIVE */
    .info-card {
        background: linear-gradient(135deg, #e0f2fe 0%, #bae6fd 100%);
        border-radius: 12px;
        padding: 1rem;
        margin-bottom: 1rem;
        border-left: 4px solid #0284c7;
    }
    
    @media (min-width: 768px) {
        .info-card {
            padding: 1.25rem 1.5rem;
            border-radius: 16px;
            margin-bottom: 1.5rem;
        }
    }
    
    .info-card h4 {
        margin: 0 0 0.3rem 0;
        color: #0c4a6e;
        font-size: 0.9rem;
    }
    
    @media (min-width: 768px) {
        .info-card h4 {
            font-size: 1rem;
            margin: 0 0 0.5rem 0;
        }
    }
    
    .info-card p {
        margin: 0;
        color: #0369a1;
        font-size: 0.85rem;
        line-height: 1.4;
    }
    
    @media (min-width: 768px) {
        .info-card p {
            font-size: 0.95rem;
            line-height: 1.5;
        }
    }
    
    /* Section title for sample data */
    .section-title {
        font-size: 1rem;
        font-weight: 700;
        color: #1e293b;
        margin: 1rem 0 0.75rem 0;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid #667eea;
    }
    
    @media (min-width: 768px) {
        .section-title {
            font-size: 1.2rem;
            margin: 1.5rem 0 1rem 0;
        }
    }
    
    /* CTA Box - MOBILE RESPONSIVE */
    .cta-box {
        background: linear-gradient(135deg, #667eea10 0%, #764ba215 100%);
        padding: 1.25rem 1rem;
        border-radius: 16px;
        text-align: center;
        border: 2px solid #667eea;
        margin-top: 1.5rem;
        box-shadow: 0 8px 30px rgba(102, 126, 234, 0.15);
    }
    
    @media (min-width: 768px) {
        .cta-box {
            padding: 2rem 2.5rem;
            border-radius: 20px;
            margin-top: 2rem;
        }
    }
    
    .cta-box h3 {
        color: #667eea;
        margin-bottom: 0.75rem;
        font-size: 1.1rem;
    }
    
    @media (min-width: 768px) {
        .cta-box h3 {
            font-size: 1.4rem;
            margin-bottom: 1rem;
        }
    }
    
    .cta-button {
        display: inline-block;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white !important;
        padding: 0.75rem 1.5rem;
        border-radius: 25px;
        text-decoration: none;
        font-weight: bold;
        font-size: 0.95rem;
        margin-top: 0.75rem;
        box-shadow: 0 4px 20px rgba(102, 126, 234, 0.4);
    }
    
    @media (min-width: 768px) {
        .cta-button {
            padding: 1rem 2.5rem;
            font-size: 1.1rem;
            margin-top: 1rem;
        }
    }
    
    .cta-button:hover {
        transform: translateY(-3px);
        box-shadow: 0 8px 30px rgba(102, 126, 234, 0.5);
    }
    
    /* Hide file uploader completely */
    .stFileUploader {
        display: none !important;
    }
    
    /* Scrollable result table with visible scrollbar */
    .stDataFrame {
        max-height: 500px;
        overflow-y: auto !important;
        overflow-x: auto !important;
    }
    
    .stDataFrame::-webkit-scrollbar {
        width: 10px;
        height: 10px;
    }
    
    .stDataFrame::-webkit-scrollbar-track {
        background: #f1f1f1;
        border-radius: 5px;
    }
    
    .stDataFrame::-webkit-scrollbar-thumb {
        background: #888;
        border-radius: 5px;
    }
    
    .stDataFrame::-webkit-scrollbar-thumb:hover {
        background: #555;
    }
    
    /* Welcome box - MOBILE RESPONSIVE */
    .welcome-box {
        text-align: center;
        padding: 2rem 1rem;
        background: linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%);
        border-radius: 16px;
        margin: 1rem 0;
    }
    
    @media (min-width: 768px) {
        .welcome-box {
            padding: 3rem 2rem;
            margin: 2rem 0;
        }
    }
    
    .welcome-box .welcome-icon {
        font-size: 3rem;
        margin-bottom: 0.5rem;
    }
    
    .welcome-box h2 {
        color: #1e293b;
        font-size: 1.1rem;
        margin: 0 0 0.5rem 0;
    }
    
    @media (min-width: 768px) {
        .welcome-box h2 {
            font-size: 1.4rem;
        }
    }
    
    .welcome-box p {
        color: #64748b;
        font-size: 0.9rem;
        margin: 0;
    }
    
    /* Language selector styling */
    .lang-selector {
        display: flex;
        justify-content: center;
        gap: 0.5rem;
        margin-bottom: 1rem;
    }
</style>
""", unsafe_allow_html=True)

# ============================================================================
# SINGLE PAGE LAYOUT - NO SIDEBAR
# ============================================================================

# Language selection - IN PAGE (not sidebar)
lang = st.radio(
    "🌐 Language / Langue",
    ["English", "Français"],
    horizontal=True,
    label_visibility="collapsed"
)
is_fr = lang == "Français"

# Translations
t = {
    "title": "AynalyxAI Demo" if not is_fr else "Démo AynalyxAI",
    "subtitle": "AI-Powered Anomaly Detection" if not is_fr else "Détection d'Anomalies par IA",
    "demo_notice": "🎯 FREE DEMO — Try with our sample data below!" if not is_fr else "🎯 DÉMO GRATUITE — Essayez avec nos données d'exemple ci-dessous!",
    "sample_title": "📊 Select Sample Data" if not is_fr else "📊 Sélectionner des Données",
    "sample_invoices": "📄 Invoices" if not is_fr else "📄 Factures",
    "sample_expenses": "💰 Expenses" if not is_fr else "💰 Dépenses",
    "sample_payroll": "👥 Payroll" if not is_fr else "👥 Paie",
    "sample_inventory": "📦 Inventory" if not is_fr else "📦 Inventaire",
    "run_analysis": "🔍 Run Anomaly Detection" if not is_fr else "🔍 Lancer la Détection",
    "results": "🎯 Detection Results" if not is_fr else "🎯 Résultats de Détection",
    "anomalies_found": "anomalies detected" if not is_fr else "anomalies détectées",
    "download_excel": "📥 Download Results (Excel)" if not is_fr else "📥 Télécharger (Excel)",
    "data_preview": "📋 Data Preview" if not is_fr else "📋 Aperçu des Données",
    "critical": "CRITICAL" if not is_fr else "CRITIQUE",
    "high": "HIGH" if not is_fr else "ÉLEVÉ",
    "medium": "MEDIUM" if not is_fr else "MOYEN", 
    "low": "LOW" if not is_fr else "FAIBLE",
    "normal": "Normal" if not is_fr else "Normal",
    "explanation": "Explanation" if not is_fr else "Explication",
    "level": "Level" if not is_fr else "Niveau",
    "ai_score": "AI Score" if not is_fr else "Score IA",
    "deviation_score": "Deviation Score" if not is_fr else "Score Déviation",
    "welcome_title": "� Select Sample Data to Begin" if not is_fr else "� Sélectionnez des Données pour Commencer",
    "welcome_text": "Choose one of the 4 sample datasets below to see AI anomaly detection in action." if not is_fr else "Choisissez l'un des 4 jeux de données ci-dessous pour voir la détection d'anomalies en action.",
    "get_full": "Get Full Desktop Version" if not is_fr else "Obtenir la Version Complète",
    "full_features": """✅ Upload your own CSV/Excel files
✅ AI-powered anomaly detection (Isolation Forest + Z-score)
✅ Data cleaning & validation wizard
✅ Smart grouping & aggregation
✅ Custom ratio calculations
✅ Multi-language (English/French)
✅ Professional Excel exports with formatting
✅ 100% offline — your data stays private
✅ No subscription — one-time purchase""" if not is_fr else """✅ Importez vos propres fichiers CSV/Excel
✅ Détection d'anomalies par IA (Isolation Forest + Z-score)
✅ Assistant de nettoyage des données
✅ Agrégation et regroupement intelligent
✅ Calculs de ratios personnalisés
✅ Bilingue (Français/Anglais)
✅ Exports Excel professionnels formatés
✅ 100% hors-ligne — vos données restent privées
✅ Achat unique — pas d'abonnement""",
    "security_title": "⚠️ Security Notice" if not is_fr else "⚠️ Avis de Sécurité",
    "security_text": "**Windows may show a security warning** when you download and run the app. This is completely normal for new software from independent developers. AynalyxAI is safe and built with standard open-source tools. To proceed: click **'More info'** → **'Run anyway'**. Your antivirus may also scan the file — this is normal." if not is_fr else "**Windows peut afficher un avertissement de sécurité** lorsque vous téléchargez et lancez l'application. C'est tout à fait normal pour les nouveaux logiciels de développeurs indépendants. AynalyxAI est sécuritaire et construit avec des outils open-source standards. Pour continuer : cliquez sur **« Plus d'infos »** → **« Exécuter quand même »**. Votre antivirus peut aussi scanner le fichier — c'est normal.",
    # Aggregation/Grouping translations
    "group_by": "📊 Group by (Optional)" if not is_fr else "📊 Grouper par (Optionnel)",
    "no_grouping": "No grouping (analyze all rows)" if not is_fr else "Pas de regroupement (analyser toutes les lignes)",
    "aggregated_results": "Aggregated Results by" if not is_fr else "Résultats agrégés par",
    "run_analysis": "🔍 Run Anomaly Detection" if not is_fr else "🔍 Lancer la Détection",
}

# ============================================================================
# PROFESSIONAL ANOMALY DETECTION - Matching main app (modeling.py)
# ============================================================================
# Uses Isolation Forest (negative scores for anomalies) + Direction-aware Z-scores
# Hybrid classification: Both AI + Statistical methods combined

def clean_column_name(col_name: str) -> str:
    """Convert technical column names to user-friendly format."""
    return col_name.replace('_', ' ')

def run_professional_anomaly_detection(
    df: pd.DataFrame,
    numeric_cols: list,
    feature_directions: dict = None,
    threshold_high: float = 2.0,
    threshold_medium: float = 1.2,
    threshold_high_anomaly_score: float = -0.15,
    threshold_medium_anomaly_score: float = -0.05,
    contamination: float = 0.05,
    is_fr: bool = False
) -> pd.DataFrame:
    """
    Professional anomaly detection matching main app logic.
    
    Returns DataFrame with:
    - Isolation_Score: Negative for anomalies (from Isolation Forest)
    - Average_Deviation: Direction-aware Z-score composite
    - Anomaly_Level: Critical/High/Medium/Low/Normal
    - Anomaly_Explanation: Human-readable explanation
    """
    df_out = df.copy()
    
    # Default: all features look for HIGH values as unusual
    if feature_directions is None:
        feature_directions = {col: "high" for col in numeric_cols}
    
    X = df_out[numeric_cols].astype(float).copy()
    
    # ========================================================================
    # ISOLATION FOREST - AI Pattern Detection
    # ========================================================================
    iso = IsolationForest(
        n_estimators=120,
        contamination=contamination,
        random_state=42,
    )
    iso.fit(X)
    scores = iso.decision_function(X)  # More NEGATIVE = More anomalous
    df_out["Isolation_Score"] = np.round(scores, 2)
    
    # ========================================================================
    # DIRECTION-AWARE Z-SCORE CALCULATION
    # ========================================================================
    scaler = StandardScaler()
    z_vals = scaler.fit_transform(X)
    
    adjusted_z = np.zeros_like(z_vals)
    col_names = list(numeric_cols)
    
    for i, col in enumerate(col_names):
        z = z_vals[:, i]
        direction = feature_directions.get(col, "high")
        
        if direction == "high":
            # Only positive deviations (above mean) are unusual
            adj = np.where(z > 0, z, 0.0)
        elif direction == "low":
            # Only negative deviations (below mean) are unusual
            adj = np.where(z < 0, -z, 0.0)
        else:  # "both"
            adj = np.abs(z)
        
        adjusted_z[:, i] = adj
        df_out[f"Deviation_{col}"] = np.round(z, 2)  # Store raw z-scores rounded
    
    # Composite: average of direction-aware deviations
    composite = np.mean(adjusted_z, axis=1)
    df_out["Average_Deviation"] = np.round(composite, 2)
    
    # ========================================================================
    # HYBRID ANOMALY CLASSIFICATION (Matching modeling.py)
    # ========================================================================
    def classify_anomaly_level(row):
        z_comp = row["Average_Deviation"]
        anom_score = row["Isolation_Score"]
        
        # Extreme thresholds (1.5x stricter)
        extreme_threshold_high_stat = threshold_high * 1.5
        extreme_threshold_high_ai = threshold_high_anomaly_score * 1.5
        
        # Minimum deviation threshold to consider AI score
        min_deviation_threshold = 0.3
        
        if z_comp < min_deviation_threshold:
            return "Normal" if not is_fr else "Normal"
        
        # CRITICAL: Both agree OR either shows extreme evidence
        both_agree_high = (z_comp >= threshold_high and 
                          anom_score <= threshold_high_anomaly_score)
        stat_extreme = z_comp >= extreme_threshold_high_stat
        ai_extreme = anom_score <= extreme_threshold_high_ai and z_comp >= threshold_medium
        
        if both_agree_high or stat_extreme or ai_extreme:
            return "CRITICAL" if not is_fr else "CRITIQUE"
        
        # HIGH: Strong evidence from one metric
        elif z_comp >= threshold_high or anom_score <= threshold_high_anomaly_score:
            return "HIGH" if not is_fr else "ÉLEVÉ"
        
        # MEDIUM: Moderate evidence
        elif (z_comp >= threshold_medium or 
              (anom_score <= threshold_medium_anomaly_score and z_comp >= min_deviation_threshold)):
            return "MEDIUM" if not is_fr else "MOYEN"
        
        # LOW: Slight deviation but noteworthy
        elif z_comp >= 0.8:
            return "LOW" if not is_fr else "FAIBLE"
        
        # NORMAL
        else:
            return "Normal" if not is_fr else "Normal"
    
    df_out["Anomaly_Level"] = df_out.apply(classify_anomaly_level, axis=1)
    
    # ========================================================================
    # GENERATE HUMAN-READABLE EXPLANATIONS
    # ========================================================================
    # Pre-compute column means for ratio calculation
    col_means = {col: X[col].mean() for col in col_names}
    
    def generate_explanation(row):
        level = row["Anomaly_Level"]
        
        if level in ["Normal", "Normal"]:
            return "Normal range" if not is_fr else "Plage normale"
        
        # Helper function to format ratio explanation
        def format_ratio_text(ratio, is_french=False):
            if ratio >= 1:
                if is_french:
                    return f"{ratio:.1f}x au-dessus moy"
                return f"{ratio:.1f}x above avg"
            elif ratio > 0:
                below_pct = (1 - ratio) * 100
                if is_french:
                    return f"{below_pct:.0f}% en-dessous moy"
                return f"{below_pct:.0f}% below avg"
            else:
                return "near zero" if not is_french else "près de zéro"
        
        # Collect contributors from each feature
        contributors = []
        for i, col in enumerate(col_names):
            raw_z = row[f"Deviation_{col}"]
            adj_z = adjusted_z[row.name, i]
            direction = feature_directions.get(col, "high")
            
            # Calculate actual ratio (value / mean)
            actual_value = row[col]
            mean_value = col_means.get(col, 1)
            if mean_value != 0:
                ratio = actual_value / mean_value
            else:
                ratio = 0
            
            if adj_z > 0.5:  # Only meaningful contributors
                contributors.append({
                    'col': col,
                    'adj_z': adj_z,
                    'raw_z': raw_z,
                    'ratio': ratio,
                    'direction': direction
                })
        
        contributors.sort(key=lambda x: x['adj_z'], reverse=True)
        
        if len(contributors) == 0:
            return "AI pattern detected" if not is_fr else "Motif IA détecté"
        
        # Build explanation
        parts = []
        top_n = min(3 if level in ["CRITICAL", "CRITIQUE"] else 2, len(contributors))
        
        for contrib in contributors[:top_n]:
            col_clean = clean_column_name(contrib['col'])
            ratio = contrib['ratio']
            parts.append(f"{col_clean} = {format_ratio_text(ratio, is_fr)}")
        
        return ", ".join(parts)
        
        return ", ".join(parts)
    
    df_out["Anomaly_Explanation"] = df_out.apply(generate_explanation, axis=1)
    
    return df_out, adjusted_z

# Sample data generators with realistic anomalies
def generate_sample_invoices():
    np.random.seed(42)
    n = 100
    
    # Normal data
    amounts = np.random.normal(1500, 400, n-8)
    quantities = np.random.randint(1, 25, n-8)
    
    # Add clear anomalies
    anomaly_amounts = [18500, 22000, 15, 8, 19800, 25, 16500, 12]  # High and low anomalies
    anomaly_quantities = [180, 250, 1, 1, 200, 1, 150, 1]
    
    all_amounts = np.concatenate([amounts, anomaly_amounts])
    all_quantities = np.concatenate([quantities, anomaly_quantities])
    
    data = {
        'Invoice_ID': [f'INV-{1000+i}' for i in range(n)],
        'Client': np.random.choice(['Acme Corp', 'TechStart Inc', 'Global Services', 'Local Shop', 'Big Enterprise', 'Quick Mart', 'Pro Solutions'], n),
        'Amount': all_amounts,
        'Quantity': all_quantities,
        'Date': pd.date_range('2024-01-01', periods=n, freq='D').strftime('%Y-%m-%d').tolist()
    }
    return pd.DataFrame(data)

def generate_sample_expenses():
    np.random.seed(43)
    n = 80
    
    amounts = np.random.normal(280, 120, n-6)
    anomaly_amounts = [5800, 7200, 5, 3, 6500, 8]  # Anomalies
    all_amounts = np.concatenate([amounts, anomaly_amounts])
    
    data = {
        'Expense_ID': [f'EXP-{2000+i}' for i in range(n)],
        'Category': np.random.choice(['Travel', 'Office Supplies', 'Software', 'Marketing', 'Utilities', 'Meals'], n),
        'Vendor': np.random.choice(['Amazon', 'Office Depot', 'Google Ads', 'Airlines Inc', 'Power Co', 'Staples'], n),
        'Amount': all_amounts,
        'Date': pd.date_range('2024-01-01', periods=n, freq='D').strftime('%Y-%m-%d').tolist()
    }
    return pd.DataFrame(data)

def generate_sample_payroll():
    np.random.seed(44)
    n = 50
    
    salaries = np.random.normal(5200, 800, n-5)
    hours = np.random.normal(160, 12, n-5)
    bonuses = np.random.normal(400, 150, n-5)
    
    # Anomalies
    anomaly_salaries = [28000, 32000, 450, 380, 26000]
    anomaly_hours = [280, 310, 35, 42, 260]
    anomaly_bonuses = [8500, 12000, 0, 0, 9000]
    
    data = {
        'Employee_ID': [f'EMP-{100+i}' for i in range(n)],
        'Department': np.random.choice(['Sales', 'Engineering', 'HR', 'Marketing', 'Finance', 'Operations'], n),
        'Salary': np.concatenate([salaries, anomaly_salaries]),
        'Hours_Worked': np.concatenate([hours, anomaly_hours]),
        'Bonus': np.concatenate([bonuses, anomaly_bonuses])
    }
    return pd.DataFrame(data)

def generate_sample_inventory():
    np.random.seed(45)
    n = 60
    
    quantities = np.random.randint(80, 400, n-6)
    unit_costs = np.random.normal(28, 8, n-6)
    
    # Anomalies
    anomaly_qty = [5500, 8200, 2, 1, 6800, 3]
    anomaly_cost = [380, 520, 2, 1, 450, 3]
    
    data = {
        'SKU': [f'SKU-{3000+i}' for i in range(n)],
        'Product': np.random.choice(['Widget A', 'Gadget B', 'Tool C', 'Part D', 'Supply E', 'Component F'], n),
        'Quantity': np.concatenate([quantities, anomaly_qty]),
        'Unit_Cost': np.concatenate([unit_costs, anomaly_cost]),
    }
    df = pd.DataFrame(data)
    df['Total_Value'] = df['Quantity'] * df['Unit_Cost']
    return df

# ============================================================================
# AGGREGATION FUNCTION - Group data and run detection on aggregates
# ============================================================================
def run_aggregation_analysis(
    df: pd.DataFrame,
    group_col: str,
    numeric_cols: list,
    is_fr: bool = False
) -> pd.DataFrame:
    """
    Aggregate data by a categorical column, then run anomaly detection on aggregated results.
    This helps find anomalies at a higher level (e.g., by Category, Vendor, Department).
    """
    # Create aggregation dictionary (sum for numeric columns)
    agg_dict = {col: 'sum' for col in numeric_cols}
    
    # Group and aggregate
    agg_df = df.groupby(group_col).agg(agg_dict).reset_index()
    
    # Add count column
    agg_df['Count'] = df.groupby(group_col).size().values
    
    # Include Count in numeric columns for anomaly detection
    agg_numeric_cols = numeric_cols + ['Count']
    
    # Run anomaly detection on aggregated data
    results, adjusted_z = run_professional_anomaly_detection(
        df=agg_df,
        numeric_cols=agg_numeric_cols,
        feature_directions=None,
        threshold_high=2.0,
        threshold_medium=1.2,
        threshold_high_anomaly_score=-0.15,
        threshold_medium_anomaly_score=-0.05,
        contamination=0.1,  # Higher contamination for smaller aggregated datasets
        is_fr=is_fr
    )
    
    return results, adjusted_z

# Header
st.markdown(f"""
<div class="main-header">
    <h1>🔬 {t['title']}</h1>
    <p>{t['subtitle']}</p>
</div>
""", unsafe_allow_html=True)

# Demo notice
st.markdown(f'<div class="demo-badge">{t["demo_notice"]}</div>', unsafe_allow_html=True)

# What is AynalyxAI info card
what_is_title = "💡 Qu'est-ce qu'AynalyxAI?" if is_fr else "💡 What is AynalyxAI?"
what_is_text = "AynalyxAI est un outil d'analyse financière intelligent qui détecte automatiquement les anomalies, erreurs et irrégularités dans vos données comptables grâce à des algorithmes d'IA avancés." if is_fr else "AynalyxAI is an intelligent financial analysis tool that automatically detects anomalies, errors, and irregularities in your accounting data using advanced AI algorithms."
st.markdown(f"""
<div class="info-card">
    <h4>{what_is_title}</h4>
    <p>{what_is_text}</p>
</div>
""", unsafe_allow_html=True)

# How It Works section with aggregation step
how_title = "📋 Comment ça Fonctionne" if is_fr else "📋 How It Works"
step1 = "Importer Excel/CSV" if is_fr else "Upload Excel/CSV"
step2 = "L'IA analyse les motifs" if is_fr else "AI analyzes patterns"
step3 = "Agrégation & Regroupement" if is_fr else "Aggregation & Grouping"
step4 = "Anomalies signalées" if is_fr else "Anomalies flagged"
step5 = "Exporter rapports" if is_fr else "Export reports"

st.markdown(f"""
<div class="how-it-works">
    <h3>{how_title}</h3>
    <div class="steps-container">
        <div class="step-box">
            <div class="step-num">1</div>
            <div class="step-text">📁 {step1}</div>
        </div>
        <div class="step-box">
            <div class="step-num">2</div>
            <div class="step-text">🤖 {step2}</div>
        </div>
        <div class="step-box">
            <div class="step-num">3</div>
            <div class="step-text">📊 {step3}</div>
        </div>
        <div class="step-box">
            <div class="step-num">4</div>
            <div class="step-text">🚨 {step4}</div>
        </div>
        <div class="step-box">
            <div class="step-num">5</div>
            <div class="step-text">📥 {step5}</div>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

# ============================================================================
# SAMPLE DATA SELECTION - IN PAGE (not sidebar)
# ============================================================================
st.markdown(f'<p class="section-title">{t["sample_title"]}</p>', unsafe_allow_html=True)

# 2x2 grid of sample buttons (works great on mobile)
col1, col2, col3, col4 = st.columns(4)

with col1:
    btn_invoices = st.button(t['sample_invoices'], use_container_width=True, type="primary")
with col2:
    btn_expenses = st.button(t['sample_expenses'], use_container_width=True, type="primary")
with col3:
    btn_payroll = st.button(t['sample_payroll'], use_container_width=True, type="primary")
with col4:
    btn_inventory = st.button(t['sample_inventory'], use_container_width=True, type="primary")

# Track which sample is selected
if btn_invoices:
    st.session_state['sample_type'] = 'invoices'
elif btn_expenses:
    st.session_state['sample_type'] = 'expenses'
elif btn_payroll:
    st.session_state['sample_type'] = 'payroll'
elif btn_inventory:
    st.session_state['sample_type'] = 'inventory'

# Load data based on selection
df = None
data_name = ""
if 'sample_type' in st.session_state:
    sample = st.session_state['sample_type']
    if sample == 'invoices':
        df = generate_sample_invoices()
        data_name = "Invoices" if not is_fr else "Factures"
    elif sample == 'expenses':
        df = generate_sample_expenses()
        data_name = "Expenses" if not is_fr else "Dépenses"
    elif sample == 'payroll':
        df = generate_sample_payroll()
        data_name = "Payroll" if not is_fr else "Paie"
    elif sample == 'inventory':
        df = generate_sample_inventory()
        data_name = "Inventory" if not is_fr else "Inventaire"

# Main content
if df is not None:
    # Data preview
    st.subheader(f"{t['data_preview']} — {data_name}")
    st.dataframe(df.head(10), use_container_width=True, height=300)
    st.caption(f"{'Affichage de 10 sur' if is_fr else 'Showing 10 of'} {len(df)} {'lignes' if is_fr else 'rows'}")
    
    # Get numeric and categorical columns
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    categorical_cols = df.select_dtypes(include=['object']).columns.tolist()
    
    # ========================================================================
    # AGGREGATION/GROUPING OPTION
    # ========================================================================
    st.markdown("---")
    
    # Create grouping options
    group_options = [t['no_grouping']] + categorical_cols
    
    # Two-column layout: dropdown + button
    col_group, col_btn = st.columns([3, 1])
    
    with col_group:
        group_col = st.selectbox(
            t['group_by'],
            options=group_options,
            help="Group data by a category to detect anomalies at an aggregated level (e.g., find which Department or Vendor has unusual totals)" if not is_fr else "Regroupez les données par catégorie pour détecter les anomalies au niveau agrégé (ex: trouver quel Département ou Fournisseur a des totaux inhabituels)"
        )
    
    with col_btn:
        st.markdown("<br>", unsafe_allow_html=True)  # Spacer to align button
        run_btn = st.button(t['run_analysis'], type="primary", use_container_width=True)
    
    # Determine if using aggregation
    use_aggregation = group_col != t['no_grouping']
    
    # Run analysis ONLY when button is clicked
    if run_btn and numeric_cols:
        
        st.markdown("---")
        st.subheader(t['results'])
        
        # Show aggregation info if grouping is selected
        if use_aggregation:
            st.info(f"📊 **{t['aggregated_results']} {group_col}** — {'Les données sont agrégées avant la détection' if is_fr else 'Data is aggregated before detection'}")
        
        with st.spinner("🔍 " + ("Analyse en cours..." if is_fr else "Analyzing...")):
            # ================================================================
            # PROFESSIONAL ANOMALY DETECTION (matching modeling.py)
            # Uses Isolation Forest (negative scores) + Direction-aware Z-scores
            # Supports both raw data analysis AND aggregated/grouped analysis
            # ================================================================
            
            if use_aggregation:
                # Run aggregated analysis
                results, adjusted_z_matrix = run_aggregation_analysis(
                    df=df,
                    group_col=group_col,
                    numeric_cols=numeric_cols,
                    is_fr=is_fr
                )
            else:
                # Run standard analysis on all rows
                results, adjusted_z_matrix = run_professional_anomaly_detection(
                    df=df,
                    numeric_cols=numeric_cols,
                    feature_directions=None,  # Default: high values are unusual
                    threshold_high=2.0,
                    threshold_medium=1.2,
                    threshold_high_anomaly_score=-0.15,
                    threshold_medium_anomaly_score=-0.05,
                    contamination=0.05,
                    is_fr=is_fr
                )
            
            # Sort by Isolation_Score (most negative = most anomalous first)
            results_sorted = results.sort_values('Isolation_Score', ascending=True)
            
            # Count anomalies by level (classification uses translated labels already)
            n_critical = len(results_sorted[results_sorted['Anomaly_Level'] == t['critical']])
            n_high = len(results_sorted[results_sorted['Anomaly_Level'] == t['high']])
            n_medium = len(results_sorted[results_sorted['Anomaly_Level'] == t['medium']])
            n_low = len(results_sorted[results_sorted['Anomaly_Level'] == t['low']])
            n_normal = len(results_sorted[results_sorted['Anomaly_Level'] == t['normal']])
            n_total = n_critical + n_high + n_medium + n_low
            
            # Statistics cards - Mobile responsive (using CSS classes)
            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                st.markdown(f"""
                <div class="stat-card stat-critical">
                    <div class="stat-number" style="color: #dc2626;">{n_critical}</div>
                    <div class="stat-label">🔴 {t['critical']}</div>
                </div>
                """, unsafe_allow_html=True)
            with col2:
                st.markdown(f"""
                <div class="stat-card stat-high">
                    <div class="stat-number" style="color: #ef4444;">{n_high}</div>
                    <div class="stat-label">🟠 {t['high']}</div>
                </div>
                """, unsafe_allow_html=True)
            with col3:
                st.markdown(f"""
                <div class="stat-card stat-medium">
                    <div class="stat-number" style="color: #f59e0b;">{n_medium}</div>
                    <div class="stat-label">🟡 {t['medium']}</div>
                </div>
                """, unsafe_allow_html=True)
            with col4:
                st.markdown(f"""
                <div class="stat-card stat-low">
                    <div class="stat-number" style="color: #22c55e;">{n_low}</div>
                    <div class="stat-label">🟢 {t['low']}</div>
                </div>
                """, unsafe_allow_html=True)
            with col5:
                st.markdown(f"""
                <div class="stat-card" style="border-color: #667eea; background: linear-gradient(135deg, #fff 0%, #e0e7ff 100%);">
                    <div class="stat-number" style="color: #667eea;">{n_total}</div>
                    <div class="stat-label">📊 Total</div>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            # Show summary - different message for aggregated vs raw data
            if use_aggregation:
                row_label = "groupes" if is_fr else "groups"
                st.info(f"📊 {'Affichage de' if is_fr else 'Showing'} **{len(results_sorted)}** {row_label} ({'agrégé par' if is_fr else 'aggregated by'} {group_col}) — **{n_total}** {'anomalies' if is_fr else 'anomalies'}, **{n_normal}** {'normaux' if is_fr else 'normal'}")
            else:
                st.info(f"📊 {'Affichage de toutes les' if is_fr else 'Showing all'} **{len(results_sorted)}** {'lignes' if is_fr else 'rows'} — **{n_total}** {'anomalies détectées' if is_fr else 'anomalies detected'}, **{n_normal}** {'normales' if is_fr else 'normal'}")
            
            # Create display dataframe with ALL rows (no filtering!)
            display_df = results_sorted.reset_index(drop=True).copy()
            
            # Remove the per-column Deviation_ columns for cleaner display
            deviation_cols = [c for c in display_df.columns if c.startswith('Deviation_')]
            display_df = display_df.drop(columns=deviation_cols)
            
            # Rename columns for display (matching the new column names)
            col_rename = {
                'Anomaly_Level': t['level'],
                'Average_Deviation': t['deviation_score'],
                'Isolation_Score': t['ai_score'],
                'Anomaly_Explanation': t['explanation']
            }
            display_df = display_df.rename(columns=col_rename)
            
            # Round the scores for display (2 decimal places)
            display_df[t['deviation_score']] = display_df[t['deviation_score']].round(2)
            display_df[t['ai_score']] = display_df[t['ai_score']].round(2)
            
            # Apply color styling function
            def color_level(val):
                if val == t['critical']:
                    return 'background-color: #dc2626; color: white; font-weight: bold;'
                elif val == t['high']:
                    return 'background-color: #ef4444; color: white; font-weight: bold;'
                elif val == t['medium']:
                    return 'background-color: #f59e0b; color: white; font-weight: bold;'
                elif val == t['low']:
                    return 'background-color: #22c55e; color: white; font-weight: bold;'
                return ''
            
            def color_row(row):
                level = row[t['level']]
                if level == t['critical']:
                    return ['background-color: #fee2e2;'] * len(row)
                elif level == t['high']:
                    return ['background-color: #fef2f2;'] * len(row)
                elif level == t['medium']:
                    return ['background-color: #fffbeb;'] * len(row)
                elif level == t['low']:
                    return ['background-color: #f0fdf4;'] * len(row)
                return [''] * len(row)
            
            # Style and display ALL results - no height limit, show everything
            styled_df = display_df.style.apply(color_row, axis=1).map(
                color_level, subset=[t['level']]
            )
            
            # Display ALL rows with scrollbar (capped height)
            st.dataframe(styled_df, use_container_width=True, height=450)
            
            # Download button with formatted Excel
            st.markdown("<br>", unsafe_allow_html=True)
            
            # Create formatted Excel file
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                display_df.to_excel(writer, index=False, sheet_name='Results', startrow=0)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Results']
                
                # Define color fills for each level
                fill_critical = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
                fill_critical_row = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                fill_high = PatternFill(start_color='EF4444', end_color='EF4444', fill_type='solid')
                fill_high_row = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
                fill_medium = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
                fill_medium_row = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
                fill_low = PatternFill(start_color='22C55E', end_color='22C55E', fill_type='solid')
                fill_low_row = PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid')
                fill_header = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
                
                font_white = Font(color='FFFFFF', bold=True)
                font_header = Font(color='FFFFFF', bold=True, size=11)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Style header row
                for col_num, cell in enumerate(worksheet[1], 1):
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                
                # Find the Level column index
                level_col_idx = None
                for idx, col in enumerate(display_df.columns, 1):
                    if col == t['level']:
                        level_col_idx = idx
                        break
                
                # Apply conditional formatting to each row
                for row_num in range(2, len(display_df) + 2):
                    level_value = worksheet.cell(row=row_num, column=level_col_idx).value if level_col_idx else None
                    
                    # Determine row fill based on level
                    row_fill = None
                    level_fill = None
                    if level_value == t['critical']:
                        row_fill = fill_critical_row
                        level_fill = fill_critical
                    elif level_value == t['high']:
                        row_fill = fill_high_row
                        level_fill = fill_high
                    elif level_value == t['medium']:
                        row_fill = fill_medium_row
                        level_fill = fill_medium
                    elif level_value == t['low']:
                        row_fill = fill_low_row
                        level_fill = fill_low
                    
                    # Apply formatting to all cells in row
                    for col_num in range(1, len(display_df.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical='center')
                        
                        if row_fill:
                            cell.fill = row_fill
                        
                        # Special formatting for Level column
                        if col_num == level_col_idx and level_fill:
                            cell.fill = level_fill
                            cell.font = font_white
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Auto-fit column widths
                for col_num, column in enumerate(display_df.columns, 1):
                    max_length = len(str(column))
                    for row in worksheet.iter_rows(min_row=2, max_row=len(display_df) + 1, min_col=col_num, max_col=col_num):
                        for cell in row:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                    worksheet.column_dimensions[worksheet.cell(row=1, column=col_num).column_letter].width = min(50, max_length + 2)
                
                # Add auto-filter
                worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=1, column=len(display_df.columns)).column_letter}{len(display_df) + 1}"
                
                # Freeze header row
                worksheet.freeze_panes = 'A2'
            
            output.seek(0)
            
            st.download_button(
                label=t['download_excel'],
                data=output,
                file_name=f"aynalyxai_{data_name.lower()}_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
        
        # CTA Box - mobile responsive with language-specific Gumroad links
        gumroad_link = "https://aynalyx.gumroad.com/l/jrepve" if is_fr else "https://aynalyx.gumroad.com/l/qpqmv"
        features_html = t['full_features'].replace('\n', '<br>')
        st.markdown(f"""
        <div class="cta-box">
            <h3>🚀 {t['get_full']}</h3>
            <div style="color: #555; font-size: 0.8rem; text-align: left; display: inline-block; margin: 0.5rem auto; line-height: 1.6;">
                {features_html}
            </div>
            <br>
            <a href="{gumroad_link}" target="_blank" class="cta-button">
                💎 {'Obtenir AynalyxAI Pro' if is_fr else 'Get AynalyxAI Pro'}
            </a>
        </div>
        """, unsafe_allow_html=True)
        
        # Security Notice - mobile responsive
        st.markdown(f"""
        <div style="background: #fef3c7; border: 2px solid #f59e0b; border-radius: 10px; padding: 0.75rem 1rem; margin-top: 1rem;">
            <h4 style="color: #92400e; margin: 0 0 0.3rem 0; font-size: 0.85rem;">{t['security_title']}</h4>
            <p style="color: #78350f; margin: 0; font-size: 0.75rem; line-height: 1.4;">{t['security_text']}</p>
        </div>
        """, unsafe_allow_html=True)

else:
    # Welcome screen when no data selected - mobile responsive
    st.markdown(f"""
    <div class="welcome-box">
        <div class="welcome-icon">&#9757;</div>
        <h2>{t['welcome_title']}</h2>
        <p>{t['welcome_text']}</p>
    </div>
    """, unsafe_allow_html=True)

# Footer - with language-specific Gumroad link
footer_gumroad = "https://aynalyx.gumroad.com/l/jrepve" if is_fr else "https://aynalyx.gumroad.com/l/qpqmv"
footer_text = "Obtenir Version Pro" if is_fr else "Get Full Version"
st.markdown("---")
st.markdown(f"""
<div style="text-align: center; color: #94a3b8; padding: 0.5rem; font-size: 0.8rem;">
    <p>© 2025 Mubsira Analytics | <a href="{footer_gumroad}" target="_blank" style="color: #667eea;">{footer_text}</a></p>
</div>
""", unsafe_allow_html=True)
